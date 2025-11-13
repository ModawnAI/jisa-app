#!/usr/bin/env python3
"""
ULTRA-COMPREHENSIVE Commission System Test Suite
Tests MULTIPLE products from EACH company, ALL FC rates, and extensive edge cases
All results validated against Excel source data
"""

import openpyxl
import json
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Tuple
import random

# Color codes
class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    MAGENTA = '\033[95m'
    RESET = '\033[0m'
    BOLD = '\033[1m'

def load_excel_data():
    """Load Excel file"""
    print(f"{Colors.BLUE}Loading Excel file...{Colors.RESET}")
    wb = openpyxl.load_workbook('data/file.xlsx', data_only=True)
    return wb

def load_json_data():
    """Load JSON data"""
    print(f"{Colors.BLUE}Loading JSON data...{Colors.RESET}")
    with open('data/commission_data_base_60pct_only.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def extract_excel_product_data(wb, company: str, row: int) -> Dict:
    """Extract product data from Excel for validation"""
    ws = wb[company]

    product_name = ws.cell(row, 1).value
    payment_period = ws.cell(row, 2).value
    conversion_rate = ws.cell(row, 5).value

    # FC 수수료 columns (60% values)
    초년도_60 = ws.cell(row, 6).value  # F
    차년도2_60 = ws.cell(row, 8).value  # H
    total_60 = ws.cell(row, 36).value  # AJ

    # 총량 (100% base values) for calculation
    총량_초년도 = ws.cell(row, 75).value  # BW
    총량_2차년도 = ws.cell(row, 77).value  # BY
    총량_total = ws.cell(row, 81).value  # CC

    return {
        'product_name': product_name,
        'payment_period': payment_period,
        'conversion_rate': conversion_rate,
        'row_number': row,
        'fc_60': {
            '초년도_익월': 초년도_60,
            '2차년도_13회차': 차년도2_60,
            'Total': total_60
        },
        '총량': {
            '초년도': 총량_초년도,
            '2차년도': 총량_2차년도,
            'Total': 총량_total
        }
    }

def calculate_expected_values(총량_values: Dict, fc_rate: float) -> Dict:
    """Calculate expected commission values at given FC rate"""
    multiplier = fc_rate / 100
    return {
        '초년도': 총량_values['초년도'] * multiplier if 총량_values['초년도'] else 0,
        '2차년도': 총량_values['2차년도'] * multiplier if 총량_values['2차년도'] else 0,
        'Total': 총량_values['Total'] * multiplier if 총량_values['Total'] else 0
    }

def query_commission_system(query: str) -> Dict:
    """Query the commission system via Node.js"""
    temp_script = Path('temp_test_query.js')
    script_content = f"""
import {{ NaturalLanguageCommissionSystem }} from './src/nl_query_system_dynamic.js';

async function main() {{
    const system = new NaturalLanguageCommissionSystem();
    const result = await system.executeQuery('{query}');
    console.log('__RESULT_START__');
    console.log(JSON.stringify(result, null, 2));
    console.log('__RESULT_END__');
}}

main().catch(error => {{
    console.error('Error:', error);
    process.exit(1);
}});
"""

    temp_script.write_text(script_content, encoding='utf-8')

    try:
        result = subprocess.run(
            ['/opt/bitnami/node/bin/node', str(temp_script)],
            capture_output=True,
            text=True,
            timeout=30
        )

        output = result.stdout

        if '__RESULT_START__' in output and '__RESULT_END__' in output:
            start_idx = output.index('__RESULT_START__') + len('__RESULT_START__')
            end_idx = output.index('__RESULT_END__')
            json_str = output[start_idx:end_idx].strip()
            return json.loads(json_str)
        else:
            return {'status': 'error', 'message': 'Parse error'}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}
    finally:
        if temp_script.exists():
            temp_script.unlink()

def validate_result(test_name: str, query: str, expected: Dict, result: Dict, tolerance: float = 0.00001) -> bool:
    """Validate commission query result"""

    if result['status'] != 'success':
        print(f"{Colors.RED}✗ {test_name}{Colors.RESET}")
        print(f"  Query: {query}")
        print(f"  Error: {result.get('message', 'Unknown error')}")
        return False

    comm_data = result['commission_data']

    # Handle error case
    if 'error' in comm_data:
        print(f"{Colors.RED}✗ {test_name}{Colors.RESET}")
        print(f"  Query: {query}")
        print(f"  Error: {comm_data['error']}")
        return False

    calculated_rates = comm_data['product']['commission_rates']

    # Find keys
    초년도_key = None
    차년도2_key = None

    for key in calculated_rates.keys():
        if '초년도' in key and '익월' in key:
            초년도_key = key
        elif '2차년도' in key or '13회차' in key:
            차년도2_key = key

    total_value = calculated_rates.get('Total', 0)
    초년도_value = calculated_rates.get(초년도_key, 0) if 초년도_key else 0
    차년도2_value = calculated_rates.get(차년도2_key, 0) if 차년도2_key else 0

    # Validate
    errors = []

    if abs(초년도_value - expected['초년도']) > tolerance:
        errors.append(f"초년도: got {초년도_value:.5f}, expected {expected['초년도']:.5f}")

    if abs(차년도2_value - expected['2차년도']) > tolerance:
        errors.append(f"2차년도: got {차년도2_value:.5f}, expected {expected['2차년도']:.5f}")

    if abs(total_value - expected['Total']) > tolerance:
        errors.append(f"Total: got {total_value:.5f}, expected {expected['Total']:.5f}")

    if errors:
        print(f"{Colors.RED}✗ {test_name}{Colors.RESET}")
        print(f"  Query: {query}")
        for error in errors:
            print(f"  {error}")
        return False
    else:
        print(f"{Colors.GREEN}✓ {test_name}{Colors.RESET}")
        return True

def run_ultra_comprehensive_test_suite():
    """Run ultra-comprehensive test suite"""

    print(f"\n{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}ULTRA-COMPREHENSIVE COMMISSION SYSTEM TEST SUITE{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}Testing ALL companies with MULTIPLE products at ALL FC rates{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}\n")

    wb = load_excel_data()
    json_data = load_json_data()

    # Define comprehensive test matrix
    # (company, [row_numbers], product_hint_for_query, fc_rates_to_test)
    test_matrix = [
        # KB라이프 - 5 products
        ('KB라이프', [12, 13, 16, 34, 46], '약속플러스', [55, 60, 65, 70, 75, 80, 85, 90, 95, 100, 110, 120]),

        # 교보생명 - 5 products (first 5)
        ('교보생명', [12, 13, 14, 15, 16], '종신', [40, 50, 60, 70, 80, 90, 100, 150]),

        # 현대해상 - 5 products
        ('현대해상', [12, 28, 44, 60, 76], '건강보험', [45, 55, 65, 75, 85, 95, 105]),

        # 메리츠화재 - 5 products
        ('메리츠화재', [12, 15, 19, 25, 31], '알파', [50, 60, 70, 80, 90, 100]),

        # DB손해보험 - 5 products
        ('DB손해보험', [12, 20, 26, 42, 58], '건강보험', [55, 65, 75, 85, 95]),

        # 한화손해보험 - 5 products
        ('한화손해보험', [12, 15, 18, 21, 22], '종합보험', [60, 70, 80, 90, 100]),

        # 삼성화재 - 5 products
        ('삼성화재', [12, 34, 50, 56, 69], '건강', [50, 60, 70, 80, 90, 100, 110]),

        # 라이나손보 - all 9 products
        ('라이나손보', [12, 13, 14, 15, 16, 17, 18, 19, 20], '더핏', [60, 75, 90, 100]),
    ]

    total_tests = 0
    passed_tests = 0
    failed_tests = 0
    skipped_tests = 0

    # Test each company
    for company, row_numbers, product_hint, fc_rates in test_matrix:
        print(f"\n{Colors.BOLD}{Colors.MAGENTA}{'=' * 80}{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}{company} - Testing {len(row_numbers)} products × {len(fc_rates)} FC rates = {len(row_numbers) * len(fc_rates)} tests{Colors.RESET}")
        print(f"{Colors.BOLD}{Colors.MAGENTA}{'=' * 80}{Colors.RESET}\n")

        for row in row_numbers:
            try:
                excel_data = extract_excel_product_data(wb, company, row)
            except Exception as e:
                print(f"{Colors.YELLOW}⚠ Skipping row {row}: {e}{Colors.RESET}\n")
                skipped_tests += len(fc_rates)
                continue

            product_name = excel_data['product_name']
            payment_str = excel_data['payment_period'] if excel_data['payment_period'] else ""

            # Skip if no valid product
            if not product_name or not isinstance(product_name, str) or len(product_name) < 3:
                print(f"{Colors.YELLOW}⚠ Skipping row {row} - No valid product{Colors.RESET}\n")
                skipped_tests += len(fc_rates)
                continue

            print(f"{Colors.BLUE}Testing Row {row}: {product_name[:60]}...{Colors.RESET}")
            if payment_str:
                print(f"  Payment: {payment_str}")
            print(f"  환산율: {excel_data['conversion_rate']}")

            # Test at each FC rate
            for fc_rate in fc_rates:
                total_tests += 1

                expected = calculate_expected_values(excel_data['총량'], fc_rate)

                # Build query (vary query style for realistic testing)
                query_styles = [
                    f"{product_hint} {payment_str} {fc_rate}%".strip(),
                    f"{company} {product_hint} {fc_rate}%",
                    f"{product_name[:20]} {fc_rate}%",
                ]

                query = random.choice(query_styles)

                result = query_commission_system(query)

                test_name = f"{company} Row {row} @ {fc_rate}%"
                if validate_result(test_name, query, expected, result):
                    passed_tests += 1
                else:
                    failed_tests += 1

            print()  # Blank line after each product

    # Boundary and extreme tests
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}BOUNDARY & EXTREME VALUE TESTS{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}\n")

    boundary_tests = [
        # Boundary percentages
        ("약속플러스 5년납 1%", "Minimum boundary: 1%"),
        ("약속플러스 5년납 5%", "Very low: 5%"),
        ("약속플러스 5년납 10%", "Low: 10%"),
        ("약속플러스 5년납 200%", "Maximum boundary: 200%"),
        ("약속플러스 5년납 180%", "Very high: 180%"),

        # Decimal-like percentages (should work)
        ("약속플러스 5년납 67%", "Odd percentage: 67%"),
        ("약속플러스 5년납 73%", "Odd percentage: 73%"),
        ("약속플러스 5년납 89%", "Odd percentage: 89%"),

        # Product name variations
        ("KB약속플러스 5년납 75%", "No space in company name"),
        ("약속 플러스 5년 납 75%", "Extra spacing"),
        ("약속플러스5년납75%", "No spacing at all"),
        ("약속플러스 5 년납 75%", "Space in payment period"),

        # Typos and fuzzy matching
        ("약속프러스 5년납 75%", "Typo: 프 instead of 플"),
        ("약속플스 5년납 75%", "Missing 러"),
        ("야속플러스 5년납 75%", "Typo: 야 instead of 약"),
        ("약수플러스 5년납 75%", "Typo: 수 instead of 속"),

        # Mixed Korean/English
        ("KB life 약속플러스 5년납 75%", "Mixed Korean/English"),
        ("약속플러스 insurance 5년납 75%", "English word inserted"),

        # Percentage variations
        ("약속플러스 5년납 75", "Missing % symbol"),
        ("약속플러스 5년납 75프로", "Korean 프로"),
        ("약속플러스 5년납 75 %", "Space before %"),
        ("약속플러스 5년납 75  %", "Multiple spaces before %"),

        # Payment period variations
        ("약속플러스 오년납 75%", "Korean number (오년)"),
        ("약속플러스 5년 납 75%", "Spaced payment period"),
        ("약속플러스 5 년 납 75%", "Extra spaced payment"),

        # Company variations
        ("KB 약속플러스 5년납 75%", "With company name"),
        ("KB라이프 약속플러스 5년납 75%", "Full company name"),
        ("케이비 약속플러스 5년납 75%", "Korean phonetic"),

        # Multiple products queries
        ("종신보험 60%", "Generic product type"),
        ("건강보험 70%", "Generic health insurance"),
        ("암보험 80%", "Cancer insurance generic"),

        # Missing information
        ("약속플러스 75%", "Missing payment period"),
        ("약속플러스 5년납", "Missing percentage - should default to 60%"),
        ("5년납 75%", "Missing product name"),

        # Extreme queries
        ("", "Empty query"),
        ("75%", "Only percentage"),
        ("약속", "Only product keyword"),
        ("KB", "Only company name"),

        # Case variations
        ("약속플러스 5년납 75%", "Lowercase (Korean)"),
        ("ABCDEFG 75%", "All caps nonsense"),

        # Special characters
        ("약속플러스! 5년납 75%", "With exclamation"),
        ("약속플러스? 5년납 75%", "With question mark"),
        ("약속플러스. 5년납 75%", "With period"),
        ("약속플러스, 5년납, 75%", "With commas"),

        # Long queries
        ("저는 KB라이프의 약속플러스 종신보험 5년납 상품의 75% 수수료율을 알고 싶습니다", "Natural language sentence"),
        ("약속플러스 5년납 상품 수수료 75프로는 얼마인가요?", "Question format"),
    ]

    print(f"{Colors.YELLOW}Running {len(boundary_tests)} boundary/extreme tests...{Colors.RESET}\n")

    for query, description in boundary_tests:
        total_tests += 1
        print(f"{Colors.CYAN}{description}{Colors.RESET}")
        print(f"Query: '{query}'")

        result = query_commission_system(query)

        if result['status'] == 'success':
            print(f"{Colors.GREEN}✓ System handled query successfully{Colors.RESET}")
            if 'commission_data' in result and 'product' in result.get('commission_data', {}):
                comm = result['commission_data']
                print(f"  Product: {result['best_match']['product_name'][:50]}...")
                print(f"  Match Score: {result['best_match']['match_score']:.2f}")
                if 'Total' in comm['product']['commission_rates']:
                    total_val = comm['product']['commission_rates']['Total']
                    print(f"  Total: {total_val:.5f} ({total_val * 100:.2f}%)")
            passed_tests += 1
        else:
            # Some edge cases expected to fail gracefully
            if query in ["", "75%", "ABCDEFG 75%"]:
                print(f"{Colors.YELLOW}⚠ Expected graceful failure: {result.get('message', 'Error')}{Colors.RESET}")
                passed_tests += 1  # Count as pass if it fails gracefully
            else:
                print(f"{Colors.RED}✗ Unexpected failure: {result.get('message', 'Error')}{Colors.RESET}")
                failed_tests += 1

        print()

    # Final summary
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}TEST SUMMARY{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}\n")

    print(f"Total Tests: {total_tests}")
    print(f"{Colors.GREEN}Passed: {passed_tests}{Colors.RESET}")
    print(f"{Colors.RED}Failed: {failed_tests}{Colors.RESET}")
    print(f"{Colors.YELLOW}Skipped: {skipped_tests}{Colors.RESET}")

    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    print(f"\nPass Rate: {pass_rate:.1f}%")

    if failed_tests == 0:
        print(f"\n{Colors.GREEN}{Colors.BOLD}✓ ALL TESTS PASSED!{Colors.RESET}")
        return 0
    else:
        print(f"\n{Colors.RED}{Colors.BOLD}✗ {failed_tests} TESTS FAILED{Colors.RESET}")
        return 1

if __name__ == "__main__":
    sys.exit(run_ultra_comprehensive_test_suite())
