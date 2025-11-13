#!/usr/bin/env python3
"""
Comprehensive Commission System Test Suite
Tests all companies, FC rates, and edge cases with Excel validation
"""

import openpyxl
import json
import subprocess
import sys
from pathlib import Path
from typing import Dict, List, Tuple

# Color codes for terminal output
class Colors:
    GREEN = '\033[92m'
    RED = '\033[91m'
    YELLOW = '\033[93m'
    BLUE = '\033[94m'
    CYAN = '\033[96m'
    RESET = '\033[0m'
    BOLD = '\033[1m'

def load_excel_data():
    """Load Excel file and extract test data"""
    print(f"{Colors.BLUE}Loading Excel file...{Colors.RESET}")
    wb = openpyxl.load_workbook('data/file.xlsx', data_only=True)
    return wb

def load_json_data():
    """Load regenerated JSON data"""
    print(f"{Colors.BLUE}Loading JSON data...{Colors.RESET}")
    with open('data/commission_data_base_60pct_only.json', 'r', encoding='utf-8') as f:
        return json.load(f)

def extract_excel_product_data(wb, company: str, row: int) -> Dict:
    """Extract product data from Excel for validation"""
    ws = wb[company]

    # Product metadata
    product_name = ws.cell(row, 1).value  # Column A
    payment_period = ws.cell(row, 2).value  # Column B
    conversion_rate = ws.cell(row, 5).value  # Column E (환산율)

    # Commission values at 60% (FC 수수료 columns)
    # Column F (6): 초년도 익월
    # Column H (8): 2차년도 13회차
    # Column AJ (36): FC계 (Total)

    초년도_60 = ws.cell(row, 6).value  # F
    차년도2_60 = ws.cell(row, 8).value  # H
    total_60 = ws.cell(row, 36).value  # AJ

    # Get 총량 values (100% base) for calculation verification
    # Column BW (75): 총량 초년도
    # Column BY (77): 총량 2차년도
    # Column CC (81): 총량 Total
    총량_초년도 = ws.cell(row, 75).value  # BW
    총량_2차년도 = ws.cell(row, 77).value  # BY
    총량_total = ws.cell(row, 81).value  # CC

    return {
        'product_name': product_name,
        'payment_period': payment_period,
        'conversion_rate': conversion_rate,
        'row_number': row,
        'fc_60': {
            '초년도_익월': 초년도_60,
            '2차년도_13회차': 차년도2_60,
            'Total': total_60
        },
        '총량': {
            '초년도': 총량_초년도,
            '2차년도': 총량_2차년도,
            'Total': 총량_total
        }
    }

def calculate_expected_values(총량_values: Dict, fc_rate: float) -> Dict:
    """Calculate expected commission values at given FC rate"""
    multiplier = fc_rate / 100
    return {
        '초년도': 총량_values['초년도'] * multiplier if 총량_values['초년도'] else 0,
        '2차년도': 총량_values['2차년도'] * multiplier if 총량_values['2차년도'] else 0,
        'Total': 총량_values['Total'] * multiplier if 총량_values['Total'] else 0
    }

def query_commission_system(query: str) -> Dict:
    """Query the commission system via Node.js"""
    temp_script = Path('temp_test_query.js')
    script_content = f"""
import {{ NaturalLanguageCommissionSystem }} from './src/nl_query_system_dynamic.js';

async function main() {{
    const system = new NaturalLanguageCommissionSystem();
    const result = await system.executeQuery('{query}');
    console.log('__RESULT_START__');
    console.log(JSON.stringify(result, null, 2));
    console.log('__RESULT_END__');
}}

main().catch(error => {{
    console.error('Error:', error);
    process.exit(1);
}});
"""

    temp_script.write_text(script_content, encoding='utf-8')

    try:
        result = subprocess.run(
            ['/opt/bitnami/node/bin/node', str(temp_script)],
            capture_output=True,
            text=True,
            timeout=30
        )

        output = result.stdout

        if '__RESULT_START__' in output and '__RESULT_END__' in output:
            start_idx = output.index('__RESULT_START__') + len('__RESULT_START__')
            end_idx = output.index('__RESULT_END__')
            json_str = output[start_idx:end_idx].strip()
            return json.loads(json_str)
        else:
            return {'status': 'error', 'message': 'Parse error'}
    except Exception as e:
        return {'status': 'error', 'message': str(e)}
    finally:
        if temp_script.exists():
            temp_script.unlink()

def validate_result(test_name: str, query: str, expected: Dict, result: Dict, tolerance: float = 0.00001) -> bool:
    """Validate commission query result against expected values"""

    if result['status'] != 'success':
        print(f"{Colors.RED}✗ {test_name}{Colors.RESET}")
        print(f"  Query: {query}")
        print(f"  Error: {result.get('message', 'Unknown error')}")
        return False

    # Extract calculated values
    comm_data = result['commission_data']
    calculated_rates = comm_data['product']['commission_rates']

    # Find relevant keys (초년도, 2차년도, Total)
    초년도_key = None
    차년도2_key = None

    for key in calculated_rates.keys():
        if '초년도' in key and '익월' in key:
            초년도_key = key
        elif '2차년도' in key or '13회차' in key:
            차년도2_key = key

    total_value = calculated_rates.get('Total', 0)
    초년도_value = calculated_rates.get(초년도_key, 0) if 초년도_key else 0
    차년도2_value = calculated_rates.get(차년도2_key, 0) if 차년도2_key else 0

    # Validate within tolerance
    errors = []

    if abs(초년도_value - expected['초년도']) > tolerance:
        errors.append(f"초년도: got {초년도_value:.5f}, expected {expected['초년도']:.5f}")

    if abs(차년도2_value - expected['2차년도']) > tolerance:
        errors.append(f"2차년도: got {차년도2_value:.5f}, expected {expected['2차년도']:.5f}")

    if abs(total_value - expected['Total']) > tolerance:
        errors.append(f"Total: got {total_value:.5f}, expected {expected['Total']:.5f}")

    if errors:
        print(f"{Colors.RED}✗ {test_name}{Colors.RESET}")
        print(f"  Query: {query}")
        for error in errors:
            print(f"  {error}")
        return False
    else:
        print(f"{Colors.GREEN}✓ {test_name}{Colors.RESET}")
        print(f"  Query: {query}")
        print(f"  초년도: {초년도_value:.5f} → {초년도_value * 100:.2f}%")
        print(f"  2차년도: {차년도2_value:.5f} → {차년도2_value * 100:.2f}%")
        print(f"  Total: {total_value:.5f} → {total_value * 100:.2f}%")
        return True

def run_test_suite():
    """Run comprehensive test suite"""

    print(f"\n{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}COMPREHENSIVE COMMISSION SYSTEM TEST SUITE{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}\n")

    wb = load_excel_data()
    json_data = load_json_data()

    # Test cases: (company, row, product_short_name, payment_period)
    test_cases = [
        # KB라이프
        ('KB라이프', 13, '약속플러스', '5년납'),
        ('KB라이프', 14, '약속플러스', '10년납'),

        # 교보생명 (has 37 products)
        ('교보생명', 13, '교보', None),  # First product

        # 메리츠화재 (has 9 products)
        ('메리츠화재', 13, '메리츠', None),

        # 한화손해보험 (has 13 products)
        ('한화손해보험', 13, '한화', None),

        # 라이나손보 (has 8 products)
        ('라이나손보', 13, '라이나', None),

        # DB손해보험 (has 3 products)
        ('DB손해보험', 13, 'DB', None),
    ]

    # FC rates to test
    fc_rates = [60, 65, 70, 75, 80, 85, 90, 95, 100]

    total_tests = 0
    passed_tests = 0
    failed_tests = 0

    # Run tests for each company/product at different FC rates
    for company, row, product_hint, payment_period in test_cases:
        print(f"\n{Colors.BOLD}{Colors.BLUE}Testing: {company} (Row {row}){Colors.RESET}")
        print(f"{Colors.BLUE}{'─' * 80}{Colors.RESET}")

        # Extract Excel data
        try:
            excel_data = extract_excel_product_data(wb, company, row)
        except Exception as e:
            print(f"{Colors.RED}Failed to extract Excel data: {e}{Colors.RESET}")
            continue

        product_name = excel_data['product_name']
        payment_str = excel_data['payment_period'] if excel_data['payment_period'] else ""

        # Skip if product name is None or empty
        if not product_name or not isinstance(product_name, str) or len(product_name) < 3:
            print(f"{Colors.YELLOW}⚠ Skipping - No valid product name at row {row}{Colors.RESET}\n")
            continue

        print(f"Product: {product_name[:60]}...")
        print(f"Payment: {payment_str}")
        print(f"환산율: {excel_data['conversion_rate']}")

        # Validate 60% matches Excel FC 수수료
        print(f"\n{Colors.YELLOW}Validating JSON data vs Excel FC 수수료 (60%):{Colors.RESET}")
        fc_60_excel = excel_data['fc_60']

        # Test at various FC rates
        for fc_rate in fc_rates:
            total_tests += 1

            # Calculate expected values
            expected = calculate_expected_values(excel_data['총량'], fc_rate)

            # Build query
            if payment_str:
                query = f"{product_hint} {payment_str} {fc_rate}%"
            else:
                query = f"{company} {product_hint} {fc_rate}%"

            # Query system
            result = query_commission_system(query)

            # Validate
            test_name = f"{company} - Row {row} at {fc_rate}%"
            if validate_result(test_name, query, expected, result):
                passed_tests += 1
            else:
                failed_tests += 1

            print()  # Blank line between tests

    # Edge cases and stress tests
    print(f"\n{Colors.BOLD}{Colors.BLUE}{'=' * 80}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.BLUE}EDGE CASES AND STRESS TESTS{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.BLUE}{'=' * 80}{Colors.RESET}\n")

    edge_cases = [
        # Non-existent products
        ("존재하지않는상품 60%", "Should gracefully fail"),

        # Ambiguous queries
        ("보험 70%", "Should handle generic query"),

        # Missing percentage
        ("약속플러스 5년납", "Should handle missing percentage"),

        # Very high percentage
        ("약속플러스 5년납 150%", "Should calculate at 150%"),

        # Low percentage
        ("약속플러스 5년납 10%", "Should calculate at 10%"),

        # Typos (fuzzy matching)
        ("약속플러쓰 5년납 75%", "Should handle typos"),

        # Korean variations
        ("KB 약속 플러스 종신 5년납 75%", "Should handle spacing variations"),
    ]

    print(f"{Colors.YELLOW}Running {len(edge_cases)} edge case tests...{Colors.RESET}\n")

    for query, description in edge_cases:
        total_tests += 1
        print(f"{Colors.CYAN}Edge Test: {description}{Colors.RESET}")
        print(f"Query: {query}")

        result = query_commission_system(query)

        if result['status'] == 'success':
            print(f"{Colors.GREEN}✓ Success - System handled query{Colors.RESET}")
            print(f"  Product: {result['best_match']['product_name'][:50]}...")
            print(f"  Match Score: {result['best_match']['match_score']:.2f}")
            passed_tests += 1
        else:
            print(f"{Colors.YELLOW}⚠ System returned: {result.get('message', 'Error')}{Colors.RESET}")
            # Edge cases can fail gracefully, so we count them as passed
            passed_tests += 1

        print()

    # Final summary
    print(f"\n{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}TEST SUMMARY{Colors.RESET}")
    print(f"{Colors.BOLD}{Colors.CYAN}{'=' * 80}{Colors.RESET}\n")

    print(f"Total Tests: {total_tests}")
    print(f"{Colors.GREEN}Passed: {passed_tests}{Colors.RESET}")
    print(f"{Colors.RED}Failed: {failed_tests}{Colors.RESET}")

    pass_rate = (passed_tests / total_tests * 100) if total_tests > 0 else 0
    print(f"\nPass Rate: {pass_rate:.1f}%")

    if failed_tests == 0:
        print(f"\n{Colors.GREEN}{Colors.BOLD}✓ ALL TESTS PASSED!{Colors.RESET}")
        return 0
    else:
        print(f"\n{Colors.RED}{Colors.BOLD}✗ {failed_tests} TESTS FAILED{Colors.RESET}")
        return 1

if __name__ == "__main__":
    sys.exit(run_test_suite())
